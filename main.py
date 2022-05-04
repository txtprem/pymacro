import os, os.path
import win32com.client
import sys
import openpyxl as xl
import shutil

def old():
    if os.path.exists("D:\\excle_macro_temp\\prem_macro.xlsm"):
        xl=win32com.client.Dispatch("Excel.Application")
        xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(os.path.abspath(sys.argv[1]))
        excel_name = os.path.basename(sys.argv[1])
        os.path.expanduser("D:\\excle_macro_temp\\m4.xlsm")
        try:
            # Note: Macro_1 should not be run since it was fixed to file path of Andrea
            #       real macro from 2 to 7
            # xl.Application.Run("m4.xlsm!Module1.Macro_1")

            # xl.Application.Run("m4.xlsm!Module1.Macro_2")
            xl.Application.Run("m4.xlsm!Module1.Macro_3")
            # xl.Application.Run("m4.xlsm!Module1.Macro_4")
            # xl.Application.Run("m4.xlsm!Module1.Macro_5")
            # xl.Application.Run("m4.xlsm!Module1.Macro_6")
            # xl.Application.Run("m4.xlsm!Module1.Macro_7")

            # NOT WORK BELOW!!!!
            # xl.Application.Run("D:\\excle_macro_temp\\m4.xlsm!Module1.Macro_2")
            # xl.Application.Run("D:\\excle_macro_temp\\m4.xlsm!Module1.Macro_3")
            # xl.Application.Run("D:\\excle_macro_temp\\m4.xlsm!Module1.Macro_4")
            # xl.Application.Run("D:\\excle_macro_temp\\m4.xlsm!Module1.Macro_5")
            # xl.Application.Run("D:\\excle_macro_temp\\m4.xlsm!Module1.Macro_6")
            # xl.Application.Run("D:\\excle_macro_temp\\m4.xlsm!Module1.Macro_7")
        except:
            sys.stderr.write("Error raised when run macro file, exit...")
            xl.Application.Quit()  # Comment this out if your excel script closes
            del xl
            sys.exit(1)

        wb.SaveAs(Filename="D:\\excle_macro_temp\\" + "macro_" + excel_name)
        # wb.SaveAs(Filename="D:\\excle_macro_temp\\ciao.xlsx")

        wb.Close()
        xl.Quit() # Comment this out if your excel script closes
        del xl

def prepare_new_xlsm():
    excel_name = os.path.basename(sys.argv[1]).split('.')[0]
    shutil.copy(os.path.join(work_folder, "prem_macro.xlsm"), os.path.join(work_folder, excel_name + ".xlsm"))

    filename = sys.argv[1]
    wb1 = xl.load_workbook(filename)
    ws1 = wb1.worksheets[0]

    # opening the destination excel file
    filename1 = os.path.join(work_folder, excel_name + ".xlsm")
    wb2 = xl.load_workbook(filename1, read_only=False, keep_vba=True)
    ws2 = wb2.active

    # calculate total number of rows and
    # columns in source excel file
    mr = ws1.max_row
    mc = ws1.max_column

    # copying the cell values from source
    # excel file to destination excel file
    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            # reading cell value from source excel file
            c = ws1.cell(row=i, column=j)

            # writing the read value to destination excel file
            ws2.cell(row=i, column=j).value = c.value

    # saving the destination excel file
    wb2.save(str(os.path.join(work_folder, excel_name + ".xlsm")))
    return excel_name

def main(excel_name):
    if os.path.exists(os.path.join(work_folder, excel_name + ".xlsm")):
        xl = win32com.client.Dispatch("Excel.Application")
        xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(os.path.join(work_folder, excel_name + ".xlsm"))
        # excel_name = os.path.basename(sys.argv[1])
        # os.path.expanduser("D:\\excle_macro_temp\\m4.xlsm")
        os.path.expanduser(os.path.join(work_folder, excel_name + ".xlsm"))
        try:
            # Note: Macro_1 should not be run since it was fixed to file path of Andrea
            #       real macro from 2 to 7
            # xl.Application.Run("m4.xlsm!Module1.Macro_1")

            xl.Application.Run(excel_name + ".xlsm!Module1.Macro_2")
            xl.Application.Run(excel_name + ".xlsm!Module1.Macro_3")
            xl.Application.Run(excel_name + ".xlsm!Module1.Macro_4")
            xl.Application.Run(excel_name + ".xlsm!Module1.Macro_5")
            xl.Application.Run(excel_name + ".xlsm!Module1.Macro_6")
            xl.Application.Run(excel_name + ".xlsm!Module1.Macro_7")

        except:
            sys.stderr.write("Error raised when run macro file, exit...")
            xl.Application.Quit()  # Comment this out if your excel script closes
            del xl
            sys.exit(1)

        wb.SaveAs(Filename="D:\\excle_macro_temp\\" + "macro_" + excel_name + ".xlsm")
        # wb.SaveAs(Filename="D:\\excle_macro_temp\\ciao.xlsx")

        wb.Close()
        xl.Quit()  # Comment this out if your excel script closes
        del xl


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    # old()
    # try_v()
    work_folder = "D:\\excle_macro_temp"
    excel_n = prepare_new_xlsm()
    main(excel_n)

