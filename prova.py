# import os, os.path
# import win32com.client
# import sys
#
# import pandas as pd
# writer = pd.ExcelWriter('D:\\excle_macro_temp\\m4_dev.xlsm')
# pd.read_csv('D:\\excle_macro_temp\\RelStat_txtmi00_quan202211817831.csv').to_excel(writer,'sheetx')
# # pd.read_csv('D:\\excle_macro_temp\\RelStat_txtmi00_quan202211817831.csv').to_excel(writer)
# writer.save()

import openpyxl as xl;

# opening the source excel file
filename = "D:\\excle_macro_temp\\RelStat_txtmi00_quan202211817831.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# opening the destination excel file
filename1 = "D:\\excle_macro_temp\\m4_dev.xlsm"
wb2 = xl.load_workbook(filename1, read_only=False, keep_vba=True)
ws2 = wb2.active

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source
# excel file to destination excel file
for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row=i, column=j)

        # writing the read value to destination excel file
        ws2.cell(row=i, column=j).value = c.value

# saving the destination excel file
wb2.save(str("D:\\excle_macro_temp\\m4_dev_ok.xlsm"))